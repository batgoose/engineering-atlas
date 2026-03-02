"""
Sync Madden NFL player ratings.

Sources:
  - Madden 24/25: legacy Excel files hosted on maddenratings.weebly.com
  - Madden 26+:   player pages on maddenratings.com

Usage:
    python manage.py sync_madden_ratings               # sync current Madden year
    python manage.py sync_madden_ratings --year 26
    python manage.py sync_madden_ratings --dry-run
"""

import io
import logging
import re
import unicodedata
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import date
from html import unescape
from urllib.parse import urlparse

import openpyxl
import requests
from django.db import transaction
from django.utils.text import slugify

from gridstream.models import Player, PlayerMaddenRating

from ._base import ImportBaseCommand

logger = logging.getLogger(__name__)

# Base URL pattern. The filename differs slightly per year.
_XLSX_URLS = {
    24: "https://maddenratings.weebly.com/uploads/1/4/0/9/14097292/maddennfl24fullplayerratings.xlsx",
    25: "https://maddenratings.weebly.com/uploads/1/4/0/9/14097292/madden_nfl_25_-_full_player_ratings.xlsx",
}
_XLSX_URL_TEMPLATE = (
    "https://maddenratings.weebly.com/uploads/1/4/0/9/14097292/"
    "maddennfl{year}fullplayerratings.xlsx"
)

_MADDENRATINGS_SITEMAP_INDEX = "https://www.maddenratings.com/sitemap_index.xml"
_USER_AGENT = (
    "engineering-atlas-gridstream/1.0 (+https://github.com/jbooth/engineering-atlas)"
)


def _current_madden_year(today: date | None = None) -> int:
    """
    Madden game year currently in market.

    Examples:
      - 2025-03 -> Madden 25
      - 2025-09 -> Madden 26
      - 2026-03 -> Madden 26
    """
    today = today or date.today()
    game_year = today.year - 2000
    if today.month >= 8:
        game_year += 1
    return game_year


def _normalize_text(value: str) -> str:
    value = unescape(value or "")
    value = unicodedata.normalize("NFKD", value)
    value = value.encode("ascii", "ignore").decode("ascii")
    return value


def _normalize_name(value: str) -> str:
    value = _normalize_text(value).lower()
    value = re.sub(r"[^\w\s-]", " ", value)
    value = re.sub(r"\b(jr|sr|ii|iii|iv|v)\b", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _normalize_label(value: str) -> str:
    value = _normalize_text(value).lower()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def _normalize_madden_slug(value: str) -> str:
    value = (value or "").strip().strip("/")
    value = re.sub(r"-\d+$", "", value.lower())
    value = re.sub(r"[^a-z0-9-]+", "-", value)
    return re.sub(r"-+", "-", value).strip("-")


def _extract_xml_locs(xml_text: str) -> list[str]:
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return []
    locs: list[str] = []
    for node in root.iter():
        if node.tag.endswith("loc") and node.text:
            loc = node.text.strip()
            if loc:
                locs.append(loc)
    return locs


def _position_for_match(position: str) -> str:
    pos = (position or "").upper().strip()
    aliases = {
        "HB": "RB",
        "PK": "K",
        "LE": "DE",
        "RE": "DE",
        "NT": "DT",
        "LOLB": "OLB",
        "ROLB": "OLB",
        "MLB": "LB",
        "ILB": "LB",
        "FS": "S",
        "SS": "S",
        "LT": "T",
        "RT": "T",
        "LG": "G",
        "RG": "G",
    }
    return aliases.get(pos, pos)


_DEFAULT_YEAR = _current_madden_year()

# Map from Madden column name -> model field name
_XLSX_COLUMN_MAP = {
    "Overall Rating": "overall",
    "Speed": "speed",
    "Strength": "strength",
    "Awareness": "awareness",
    "Agility": "agility",
    "Acceleration": "acceleration",
    "Tackle": "tackle",
    "Power Moves": "power_moves",
    "Finesse Moves": "finesse_moves",
    "Throw Power": "throw_power",
    "Catching": "catching",
    "Short Route Running": "route_running",
    "Run Block": "run_block",
    "Pass Block": "pass_block",
    "Hit Power": "hit_power",
    "Man Coverage": "man_coverage",
    "Zone Coverage": "zone_coverage",
}

_RADAR_LABEL_TO_FIELD = {
    "overall": "overall",
    "general": "general_rating",
    "passing": "passing_rating",
    "receiving": "receiving_rating",
    "ball carrying": "ball_carrier_rating",
    "defense": "defense_rating",
    "blocking": "blocking_rating",
    "kicking": "kicking_rating",
}

_ATTRIBUTE_LABEL_TO_FIELD = {
    "speed": "speed",
    "strength": "strength",
    "awareness": "awareness",
    "agility": "agility",
    "acceleration": "acceleration",
    "tackle": "tackle",
    "power moves": "power_moves",
    "finesse moves": "finesse_moves",
    "throw power": "throw_power",
    "catching": "catching",
    "route running short": "route_running",
    "short route running": "route_running",
    "run block": "run_block",
    "pass block": "pass_block",
    "hit power": "hit_power",
    "man coverage": "man_coverage",
    "zone coverage": "zone_coverage",
}

_TITLE_RE = re.compile(
    r"<title>\s*(?P<title>.*?)\s*</title>", re.IGNORECASE | re.DOTALL
)
_TITLE_PLAYER_RE = re.compile(
    r"^(?P<name>.+?)\s+Madden(?:\s+NFL)?\s+(?P<year>\d+)\s+Rating(?:\s+\((?P<team>[^)]*)\))?",
    re.IGNORECASE,
)
_RADAR_RE = re.compile(
    r'chartjs-radar".*?labels:\s*\[(?P<labels>.*?)\].*?data:\s*\[(?P<data>.*?)\]',
    re.IGNORECASE | re.DOTALL,
)
_QUOTED_STR_RE = re.compile(r'"([^"]+)"')
_NUMBER_RE = re.compile(r"-?\d+(?:\.\d+)?")
_ATTRIBUTE_ROW_RE = re.compile(
    r"<li[^>]*>\s*<span[^>]*>\s*([0-9][0-9,]*(?:\.[0-9]+)?)\s*</span>\s*([^<]+?)\s*</li>",
    re.IGNORECASE | re.DOTALL,
)
_POSITION_RE = re.compile(r"\(([A-Z]{1,5})\)\s+position", re.IGNORECASE)
_OVR_FALLBACK_RE = re.compile(
    r"Madden(?:\s+NFL)?\s+\d+\s+Rating\s+is\s+(\d+)",
    re.IGNORECASE,
)
_LINE_CHART_OVR_RE = re.compile(
    r'chartjs-dashboard-line-player".*?data:\s*\[\s*([0-9]+)',
    re.IGNORECASE | re.DOTALL,
)


@dataclass
class ScrapedMaddenPlayerRow:
    url: str
    name: str
    position: str
    team: str
    ratings: dict[str, int]


class Command(ImportBaseCommand):
    help = "Sync Madden NFL player ratings from Madden data sources."

    def add_arguments(self, parser):
        super().add_arguments(parser)
        parser.add_argument(
            "--year",
            type=int,
            default=_DEFAULT_YEAR,
            help=f"Madden game year to sync (default: {_DEFAULT_YEAR})",
        )
        parser.set_defaults(batch_size=500)

    def handle(self, *args, **options):
        self.batch_size = options["batch_size"]
        self.dry_run = options["dry_run"]
        madden_year = options["year"]

        if self.dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN"))

        if madden_year >= 26:
            created, updated, skipped = self._sync_from_maddenratings_site(madden_year)
        else:
            created, updated, skipped = self._sync_from_xlsx(madden_year)

        if skipped:
            self.stdout.write(
                self.style.WARNING(
                    f"  Skipped {skipped} rows (no player match or missing OVR)"
                )
            )
        self.stdout.write(
            self.style.SUCCESS(f"\nDone! {created} created, {updated} updated.")
        )

    # ---- Madden 24/25 legacy Excel path ---------------------------------

    def _sync_from_xlsx(self, madden_year: int) -> tuple[int, int, int]:
        url = _XLSX_URLS.get(madden_year, _XLSX_URL_TEMPLATE.format(year=madden_year))
        self.stdout.write(f"Syncing Madden NFL {madden_year} from:\n  {url}")

        with self.timed_operation("Downloading Excel file"):
            try:
                resp = requests.get(
                    url, timeout=60, headers={"User-Agent": _USER_AGENT}
                )
                resp.raise_for_status()
            except Exception as exc:
                self.stderr.write(self.style.ERROR(f"Download failed: {exc}"))
                return 0, 0, 0

        with self.timed_operation("Parsing Excel"):
            wb = openpyxl.load_workbook(
                io.BytesIO(resp.content), read_only=True, data_only=True
            )
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            all_rows = [
                dict(zip(headers, row))
                for row in ws.iter_rows(min_row=2, values_only=True)
                if any(cell is not None for cell in row)
            ]

        self.stdout.write(f"  {len(all_rows)} player rows parsed")
        if not all_rows:
            self.stderr.write(self.style.ERROR("No data rows found in Excel file"))
            return 0, 0, 0

        has_full_name = "Full Name" in headers
        has_split_name = "First Name" in headers and "Last Name" in headers
        if not has_full_name and not has_split_name:
            self.stderr.write(
                self.style.ERROR(
                    "Unrecognized column layout, cannot extract player names"
                )
            )
            return 0, 0, 0

        with self.timed_operation("Building player cache"):
            name_cache, slug_cache, last_pos_cache = self._build_player_caches()
        self.stdout.write(f"  {len(name_cache)} normalized player names in cache")

        created = 0
        updated = 0
        skipped = 0
        batch: list[tuple[Player, int, dict[str, int | str | None]]] = []

        for row in all_rows:
            if has_full_name:
                full_name = self.safe_str(row.get("Full Name"))
            else:
                first = self.safe_str(row.get("First Name"))
                last = self.safe_str(row.get("Last Name"))
                full_name = f"{first} {last}".strip()
            if not full_name:
                skipped += 1
                continue

            position = self.safe_str(row.get("Position"))
            team = self.safe_str(row.get("Team"))
            player = self._match_player(
                full_name=full_name,
                position=position,
                source_url="",
                name_cache=name_cache,
                slug_cache=slug_cache,
                last_pos_cache=last_pos_cache,
            )
            if not player:
                skipped += 1
                logger.debug(
                    "No Madden XLSX player match: %s (%s, %s)",
                    full_name,
                    position,
                    team,
                )
                continue

            overall = self.safe_int(row.get("Overall Rating") or row.get("Overall"))
            if overall is None:
                skipped += 1
                continue

            defaults: dict[str, int | str | None] = {
                "position_snapshot": position,
                "team_snapshot": team,
                "overall": overall,
            }
            for col_name, field_name in _XLSX_COLUMN_MAP.items():
                if col_name in row and col_name != "Overall Rating":
                    defaults[field_name] = self.safe_int(row.get(col_name))

            batch.append((player, madden_year, defaults))
            if len(batch) >= self.batch_size:
                c, u = self._flush_batch(batch)
                created += c
                updated += u
                batch = []

        if batch:
            c, u = self._flush_batch(batch)
            created += c
            updated += u

        return created, updated, skipped

    # ---- Madden 26+ web scraper path -----------------------------------

    def _sync_from_maddenratings_site(self, madden_year: int) -> tuple[int, int, int]:
        self.stdout.write(
            f"Syncing Madden NFL {madden_year} from:\n  https://www.maddenratings.com (player pages)"
        )

        with self.timed_operation("Discovering player page URLs"):
            sitemap_urls = self._fetch_maddenratings_sitemaps()
            player_urls = self._fetch_maddenratings_player_urls(sitemap_urls)
        self.stdout.write(f"  {len(player_urls)} candidate player pages")
        if not player_urls:
            self.stderr.write(
                self.style.ERROR("No player pages discovered from sitemap")
            )
            return 0, 0, 0

        with self.timed_operation("Building player cache"):
            name_cache, slug_cache, last_pos_cache = self._build_player_caches()
        self.stdout.write(f"  {len(name_cache)} normalized player names in cache")

        target_slugs = set(slug_cache.keys())
        player_urls = [
            url
            for url in player_urls
            if _normalize_madden_slug(urlparse(url).path.strip("/")) in target_slugs
        ]
        self.stdout.write(f"  {len(player_urls)} pages after slug prefilter")
        if not player_urls:
            self.stderr.write(
                self.style.ERROR("No player pages left after slug prefilter")
            )
            return 0, 0, 0

        created = 0
        updated = 0
        skipped = 0
        batch: list[tuple[Player, int, dict[str, int | str | None]]] = []

        for index, url in enumerate(player_urls, start=1):
            try:
                resp = requests.get(
                    url, timeout=45, headers={"User-Agent": _USER_AGENT}
                )
                resp.raise_for_status()
            except Exception as exc:
                skipped += 1
                logger.debug("Madden player page fetch failed for %s: %s", url, exc)
                continue

            parsed_row = self._parse_maddenratings_player_page(
                resp.text, url, madden_year
            )
            if not parsed_row:
                skipped += 1
                continue

            player = self._match_player(
                full_name=parsed_row.name,
                position=parsed_row.position,
                source_url=parsed_row.url,
                name_cache=name_cache,
                slug_cache=slug_cache,
                last_pos_cache=last_pos_cache,
            )
            if not player:
                skipped += 1
                logger.debug(
                    "No Madden page player match: %s (%s) from %s",
                    parsed_row.name,
                    parsed_row.position,
                    parsed_row.url,
                )
                continue

            defaults: dict[str, int | str | None] = {
                "position_snapshot": parsed_row.position,
                "team_snapshot": parsed_row.team,
            }
            defaults.update(parsed_row.ratings)
            if defaults.get("overall") is None:
                skipped += 1
                continue

            batch.append((player, madden_year, defaults))
            if len(batch) >= self.batch_size:
                c, u = self._flush_batch(batch)
                created += c
                updated += u
                batch = []

            if index % 250 == 0:
                self.stdout.write(f"  Processed {index}/{len(player_urls)} pages...")

        if batch:
            c, u = self._flush_batch(batch)
            created += c
            updated += u

        return created, updated, skipped

    def _fetch_maddenratings_sitemaps(self) -> list[str]:
        try:
            resp = requests.get(
                _MADDENRATINGS_SITEMAP_INDEX,
                timeout=30,
                headers={"User-Agent": _USER_AGENT},
            )
            resp.raise_for_status()
        except Exception as exc:
            self.stderr.write(self.style.ERROR(f"Sitemap index download failed: {exc}"))
            return []

        sitemap_urls = []
        for loc in _extract_xml_locs(resp.text):
            if "post-sitemap" in loc:
                sitemap_urls.append(loc)

        # Preserve order while deduplicating.
        return list(dict.fromkeys(sitemap_urls))

    def _fetch_maddenratings_player_urls(self, sitemap_urls: list[str]) -> list[str]:
        player_urls: list[str] = []
        for sitemap_url in sitemap_urls:
            try:
                resp = requests.get(
                    sitemap_url,
                    timeout=30,
                    headers={"User-Agent": _USER_AGENT},
                )
                resp.raise_for_status()
            except Exception as exc:
                logger.debug("Sitemap download failed for %s: %s", sitemap_url, exc)
                continue

            for url in _extract_xml_locs(resp.text):
                parsed = urlparse(url)
                path = parsed.path.strip("/")
                if not path or "/" in path:
                    continue
                # Keep only likely player slug pages.
                if path.startswith(
                    ("teams", "lists", "updates", "ability", "countries", "page")
                ):
                    continue
                player_urls.append(url)

        return list(dict.fromkeys(player_urls))

    def _parse_maddenratings_player_page(
        self,
        html: str,
        source_url: str,
        expected_year: int,
    ) -> ScrapedMaddenPlayerRow | None:
        title_match = _TITLE_RE.search(html)
        if not title_match:
            return None

        title = re.sub(r"\s+", " ", unescape(title_match.group("title"))).strip()
        player_title = _TITLE_PLAYER_RE.search(title)
        if not player_title:
            return None

        year = self.safe_int(player_title.group("year"))
        if year != expected_year:
            return None

        name = self.safe_str(player_title.group("name"))
        if not name:
            return None

        team = self.safe_str(player_title.group("team"))
        position_match = _POSITION_RE.search(_normalize_text(html))
        position = (
            _position_for_match(self.safe_str(position_match.group(1)))
            if position_match
            else ""
        )

        ratings = self._extract_radar_ratings(html)
        if ratings.get("overall") is None:
            ratings["overall"] = self._extract_overall_fallback(html, expected_year)

        attr_ratings = self._extract_attribute_ratings(html)
        for field, value in attr_ratings.items():
            ratings.setdefault(field, value)

        if ratings.get("overall") is None:
            return None

        return ScrapedMaddenPlayerRow(
            url=source_url,
            name=name,
            position=position,
            team=team,
            ratings=ratings,
        )

    def _extract_radar_ratings(self, html: str) -> dict[str, int]:
        match = _RADAR_RE.search(html)
        if not match:
            return {}
        labels_blob = match.group("labels")
        data_blob = match.group("data")
        labels = [_normalize_label(v) for v in _QUOTED_STR_RE.findall(labels_blob)]
        values = [self.safe_int(v) for v in _NUMBER_RE.findall(data_blob)]

        parsed: dict[str, int] = {}
        for label, value in zip(labels, values):
            if value is None:
                continue
            field_name = _RADAR_LABEL_TO_FIELD.get(label)
            if field_name:
                parsed[field_name] = value
        return parsed

    def _extract_overall_fallback(self, html: str, expected_year: int) -> int | None:
        text = _normalize_text(html)

        # Common long-form sentence: "... Madden NFL 26 Rating is 99"
        sentence_match = _OVR_FALLBACK_RE.search(text)
        if sentence_match:
            return self.safe_int(sentence_match.group(1))

        # "Ratings over the years" table row for the requested game year.
        years_row_re = re.compile(
            rf"Madden(?:\s+NFL)?\s+{expected_year}</td>\s*<td[^>]*>\s*<span[^>]*>\s*([0-9]+)\s*</span>",
            re.IGNORECASE | re.DOTALL,
        )
        years_row_match = years_row_re.search(html)
        if years_row_match:
            return self.safe_int(years_row_match.group(1))

        # Weekly movement line chart data: first datapoint is launch OVR.
        line_match = _LINE_CHART_OVR_RE.search(html)
        if line_match:
            return self.safe_int(line_match.group(1))

        return None

    def _extract_attribute_ratings(self, html: str) -> dict[str, int]:
        parsed: dict[str, int] = {}
        for raw_value, raw_label in _ATTRIBUTE_ROW_RE.findall(html):
            field_name = _ATTRIBUTE_LABEL_TO_FIELD.get(_normalize_label(raw_label))
            if not field_name:
                continue
            if field_name in parsed:
                continue
            value = self.safe_int(str(raw_value).replace(",", ""))
            if value is None:
                continue
            parsed[field_name] = value
        return parsed

    # ---- Shared matching + write path ----------------------------------

    def _build_player_caches(self):
        name_cache: dict[str, list[Player]] = {}
        slug_cache: dict[str, list[Player]] = {}
        last_pos_cache: dict[tuple[str, str], list[Player]] = {}

        def add_multi(cache: dict, key, value):
            if not key:
                return
            cache.setdefault(key, []).append(value)

        for player in Player.objects.using("nfl").all():
            name_variants = {
                _normalize_name(player.display_name),
                _normalize_name(player.short_name),
                _normalize_name(f"{player.first_name} {player.last_name}"),
            }
            for variant in name_variants:
                add_multi(name_cache, variant, player)

            slug_variants = {
                _normalize_madden_slug(slugify(player.display_name)),
                _normalize_madden_slug(
                    slugify(f"{player.first_name} {player.last_name}")
                ),
            }
            for slug_variant in slug_variants:
                add_multi(slug_cache, slug_variant, player)

            last_name = _normalize_name(player.last_name)
            if last_name and player.position:
                add_multi(
                    last_pos_cache,
                    (last_name, _position_for_match(player.position)),
                    player,
                )

        return name_cache, slug_cache, last_pos_cache

    def _match_player(
        self,
        full_name: str,
        position: str,
        source_url: str,
        name_cache: dict[str, list[Player]],
        slug_cache: dict[str, list[Player]],
        last_pos_cache: dict[tuple[str, str], list[Player]],
    ) -> Player | None:
        normalized_name = _normalize_name(full_name)
        normalized_pos = _position_for_match(position)

        candidates = self._dedupe_candidates(name_cache.get(normalized_name, []))
        match = self._pick_candidate(candidates, normalized_pos)
        if match:
            return match

        if source_url:
            slug = _normalize_madden_slug(urlparse(source_url).path.strip("/"))
            slug_candidates = self._dedupe_candidates(slug_cache.get(slug, []))
            match = self._pick_candidate(slug_candidates, normalized_pos)
            if match:
                return match

        if normalized_name and normalized_pos:
            parts = normalized_name.split()
            if parts:
                last_name = parts[-1]
                fallback = self._dedupe_candidates(
                    last_pos_cache.get((last_name, normalized_pos), [])
                )
                if len(fallback) == 1:
                    return fallback[0]

        return None

    def _pick_candidate(self, candidates: list[Player], position: str) -> Player | None:
        if not candidates:
            return None
        if len(candidates) == 1:
            return candidates[0]
        if not position:
            return None
        pos_matches = [
            p
            for p in candidates
            if _position_for_match(p.position) == position
            or _position_for_match(p.position_group) == position
        ]
        return pos_matches[0] if len(pos_matches) == 1 else None

    def _dedupe_candidates(self, players: list[Player]) -> list[Player]:
        unique: dict[int, Player] = {}
        for player in players:
            unique[player.id] = player
        return list(unique.values())

    def _flush_batch(
        self, batch: list[tuple[Player, int, dict[str, int | str | None]]]
    ):
        created = 0
        updated = 0
        if self.dry_run:
            return len(batch), 0
        with transaction.atomic(using="nfl"):
            for player, year, defaults in batch:
                _, was_created = PlayerMaddenRating.objects.using(
                    "nfl"
                ).update_or_create(
                    player=player,
                    madden_year=year,
                    defaults=defaults,
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
        return created, updated
